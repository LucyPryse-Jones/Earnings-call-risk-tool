import os
import json
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openai import OpenAI

WORKBOOK_PATH = "risk_scanner.xlsx"
INPUT_SHEET = "Input"
RISK_LOG_SHEET = "Risk Log"
INPUT_CELL = "B4"
SOURCE_LABEL = "HSBC Q1 2025 earnings call transcript"
MODEL = "gpt-4o-mini"
TABLE_NAME = None  # set this to your Excel table name if you have more than one table

PROMPT_TEMPLATE = """
You are helping a finance analyst extract risks from an earnings call or report.
Read the text below and identify the most important risks discussed.

Return ONLY valid JSON as an array of objects.
Each object must contain exactly these keys:
- Risk Title
- Category
- Description
- Why It Matters
- Evidence
- Follow-up Question
- Severity
- Source

Rules:
- Give 5 to 7 risks only.
- Category must be one of: Macro/Policy, Credit, Capital, Operational, Regulatory/Compliance.
- Severity must be one of: Low, Medium, High.
- Evidence must be a short quote or a very short paraphrase from the text.
- Source must be: {source_label}
- Use concise, analyst-style language.

Text to analyse:
{text}
""".strip()

HEADERS = [
    "Risk Title",
    "Category",
    "Description",
    "Why It Matters",
    "Evidence",
    "Follow-up Question",
    "Severity",
    "Source",
]


def find_next_row(ws):
    row = 2
    while any(ws.cell(row=row, column=col).value not in (None, "") for col in range(1, len(HEADERS) + 1)):
        row += 1
    return row


def ensure_headers(ws):
    for i, h in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=i, value=h)


def get_input_text(wb):
    ws = wb[INPUT_SHEET]
    value = ws[INPUT_CELL].value
    if not value or not str(value).strip():
        raise ValueError(f"No input text found in {INPUT_SHEET}!{INPUT_CELL}")
    return str(value).strip()


def extract_text(resp):
    try:
        return resp.output.content.text.value.strip()
    except Exception:
        try:
            return resp.output_text.strip()
        except Exception:
            return ""


def call_model(text):
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        raise EnvironmentError("Set OPENAI_API_KEY before running this script.")

    client = OpenAI(api_key=api_key)
    prompt = PROMPT_TEMPLATE.format(text=text, source_label=SOURCE_LABEL)

    resp = client.responses.create(
        model=MODEL,
        input=prompt,
        temperature=0.2,
        max_output_tokens=1500,
    )

    raw_text = extract_text(resp)
    if not raw_text:
        raise ValueError("Model returned empty output.")

    cleaned = raw_text.strip()

    if cleaned.startswith("```"):
        lines = cleaned.splitlines()
        if lines and lines[0].startswith("```"):
            lines = lines[1:]
        if lines and lines[-1].startswith("```"):
            lines = lines[:-1]
        cleaned = "\n".join(lines).strip()
        if cleaned.lower().startswith("json"):
            cleaned = cleaned[4:].strip()

    return json.loads(cleaned)


def write_risks(ws, risks):
    start_row = find_next_row(ws)
    for risk in risks:
        for col, key in enumerate(HEADERS, start=1):
            ws.cell(row=start_row, column=col, value=risk.get(key, ""))
        start_row += 1
    return start_row - 1  # last row number used


def find_target_table(ws):
    tables = list(ws.tables.values())
    if not tables:
        return None
    if TABLE_NAME:
        return ws.tables.get(TABLE_NAME)
    return tables[0]


def resize_table(ws, last_data_row):
    table = find_target_table(ws)
    if table is None:
        return False
    end_col = get_column_letter(len(HEADERS))
    table.ref = f"A1:{end_col}{last_data_row}"
    return True


def main():
    wb = load_workbook(WORKBOOK_PATH)
    input_text = get_input_text(wb)
    risk_ws = wb[RISK_LOG_SHEET]
    ensure_headers(risk_ws)
    risks = call_model(input_text)
    last_data_row = write_risks(risk_ws, risks)
    resize_table(risk_ws, last_data_row)
    wb.save(WORKBOOK_PATH)
    print("Workbook updated successfully (table-safe).")


if __name__ == "__main__":
    main()